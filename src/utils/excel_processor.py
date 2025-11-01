"""
Excel(XLSX) 파일을 토글 구조로 변환하는 유틸리티
"""
import re
import os
from typing import List, Dict

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False


class ExcelProcessor:
    """Excel 파일을 처리하여 토글 구조로 변환"""

    def __init__(self, use_template: bool = True):
        if not EXCEL_SUPPORT:
            raise ImportError("Excel 처리를 위해 openpyxl을 설치해주세요: pip install openpyxl")

        # 템플릿 사용 여부
        self.use_template = use_template

        # 템플릿 매니저 초기화
        if use_template:
            try:
                from .template_manager import TemplateManager
                self.template_manager = TemplateManager()
            except ImportError:
                self.template_manager = None
                self.use_template = False

        # 가중치 평가기 초기화
        try:
            from .weight_evaluator import WeightEvaluator
            self.evaluator = WeightEvaluator()
            self.use_weight_evaluation = True
        except ImportError:
            self.evaluator = None
            self.use_weight_evaluation = False

    def extract_data_from_excel(self, excel_path: str) -> List[Dict]:
        """
        Excel 파일에서 데이터를 추출하고 구조화된 데이터로 반환

        Args:
            excel_path: Excel 파일 경로

        Returns:
            List[Dict]: 구조화된 행 리스트
        """
        rows_data = []

        try:
            wb = load_workbook(excel_path, data_only=True)

            # 모든 시트 처리 (또는 활성 시트만)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                # 첫 번째 행이 헤더인지 확인
                headers = []
                first_row = True

                for row_idx, row in enumerate(ws.iter_rows(values_only=False), start=1):
                    # 빈 행 건너뛰기
                    cell_values = [cell.value for cell in row]
                    if all(v is None or str(v).strip() == '' for v in cell_values):
                        continue

                    # 첫 번째 행을 헤더로 사용
                    if first_row:
                        headers = [str(cell.value).strip() if cell.value else f"열{idx+1}"
                                  for idx, cell in enumerate(row)]
                        first_row = False
                        continue

                    # 행 데이터 추출
                    row_data = {}
                    is_bold = False
                    font_size = 11
                    indent_level = 0

                    for col_idx, cell in enumerate(row):
                        if col_idx < len(headers):
                            header = headers[col_idx]
                            value = cell.value

                            # 첫 번째 셀의 스타일 정보 사용
                            if col_idx == 0:
                                if cell.font and cell.font.bold:
                                    is_bold = True
                                if cell.font and cell.font.size:
                                    font_size = cell.font.size

                                # 들여쓰기 레벨 추정 (첫 번째 열의 공백으로)
                                if value and isinstance(value, str):
                                    indent_level = len(value) - len(value.lstrip())
                                    value = value.strip()

                            row_data[header] = value

                    if row_data:
                        rows_data.append({
                            "data": row_data,
                            "sheet": sheet_name,
                            "row": row_idx,
                            "is_bold": is_bold,
                            "font_size": font_size,
                            "indent_level": indent_level,
                            "headers": headers
                        })

        except Exception as e:
            print(f"Excel 파일 읽기 오류: {e}")

        return rows_data

    def detect_structure(self, rows_data: List[Dict]) -> List[Dict]:
        """
        행 데이터에서 제목, 항목, 본문을 구분

        Args:
            rows_data: 행 데이터 리스트

        Returns:
            List[Dict]: 구조화된 항목 리스트
        """
        if not rows_data:
            return []

        structured_items = []

        for row_info in rows_data:
            data = row_info["data"]

            # 첫 번째 열의 값을 주요 텍스트로 사용
            first_col = row_info["headers"][0] if row_info["headers"] else "열1"
            text = str(data.get(first_col, "")).strip()

            if not text:
                continue

            # 구조 분석
            level = self._determine_level(text, row_info)
            item_type = self._determine_type(text, row_info)

            structured_items.append({
                "text": text,
                "level": level,
                "type": item_type,
                "data": data,
                "is_bold": row_info["is_bold"],
                "sheet": row_info["sheet"]
            })

        return structured_items

    def _determine_level(self, text: str, row_info: Dict) -> int:
        """
        행의 계층 레벨 결정 (0: 루트, 1: 하위, 2: 하위의 하위...)

        Args:
            text: 텍스트
            row_info: 행 정보

        Returns:
            int: 계층 레벨
        """
        # 들여쓰기 레벨이 있으면 사용
        if row_info["indent_level"] > 0:
            return min(3, row_info["indent_level"] // 4)  # 4칸당 1레벨

        # 숫자 패턴으로 레벨 결정
        if re.match(r'^[IVX]+\.\s+', text):  # I. II. III.
            return 0
        elif re.match(r'^\d+\.\s+', text):  # 1. 2. 3.
            return 1
        elif re.match(r'^\d+\.\d+\.\s+', text):  # 1.1. 1.2.
            return 2
        elif re.match(r'^\d+\.\d+\.\d+\.\s+', text):  # 1.1.1.
            return 3
        elif re.match(r'^\(\d+\)\s+', text):  # (1) (2)
            return 2
        elif re.match(r'^[가-힣]\)\s+', text):  # 가) 나) 다)
            return 3

        # 굵기와 폰트 크기로 레벨 결정
        if row_info["is_bold"]:
            if row_info["font_size"] >= 14:
                return 0
            elif row_info["font_size"] >= 12:
                return 1
            else:
                return 2

        return 3  # 일반 본문

    def _determine_type(self, text: str, row_info: Dict) -> str:
        """
        텍스트 타입 결정 (header, list, paragraph)

        Args:
            text: 텍스트
            row_info: 행 정보

        Returns:
            str: 타입
        """
        # 목록 패턴
        if re.match(r'^[-•·]\s+', text):
            return "list"
        elif re.match(r'^\d+[.)]\s+', text):
            return "list"
        elif re.match(r'^[가-힣][.)]\s+', text):
            return "list"

        # 짧고 굵은 텍스트는 제목
        if len(text) < 100 and row_info["is_bold"]:
            return "header"

        return "paragraph"

    def convert_to_toggle_structure(self, structured_items: List[Dict]) -> Dict:
        """
        구조화된 항목을 토글 구조로 변환

        Args:
            structured_items: 구조화된 항목 리스트

        Returns:
            Dict: 토글 구조 데이터
        """
        if not structured_items:
            return None

        # 최상위 항목 생성
        root_title = "Excel 문서"

        # 첫 번째 항목이 큰 제목이면 사용
        if structured_items and structured_items[0]["level"] == 0:
            root_title = self._clean_title(structured_items[0]["text"])
            structured_items = structured_items[1:]  # 첫 항목 제거

        root_toggle = {
            "title": root_title,
            "content": "",
            "current_score": 0,
            "max_score": 100,
            "children": [],
            "checklist": []
        }

        # 계층 구조 생성
        self._build_hierarchy(root_toggle, structured_items, 0, 0)

        return root_toggle

    def _build_hierarchy(self, parent: Dict, items: List[Dict], start_idx: int, parent_level: int) -> int:
        """
        재귀적으로 계층 구조 생성

        Args:
            parent: 부모 토글
            items: 항목 리스트
            start_idx: 시작 인덱스
            parent_level: 부모 레벨

        Returns:
            int: 처리된 마지막 인덱스
        """
        i = start_idx
        current_child = None

        while i < len(items):
            item = items[i]
            level = item["level"]
            text = item["text"]
            item_type = item["type"]

            # 같은 레벨이거나 상위 레벨이면 종료
            if level <= parent_level:
                break

            # 바로 하위 레벨인 경우
            if level == parent_level + 1:
                # 목록 항목은 체크리스트로
                if item_type == "list":
                    checklist_text = self._clean_list_text(text)
                    checklist_item = {
                        "text": checklist_text,
                        "is_checked": False,
                        "score": 1
                    }

                    # 가중치 평가 자동 적용
                    if self.use_weight_evaluation and self.evaluator:
                        weight_eval = self._auto_evaluate_checklist(checklist_text)
                        if weight_eval:
                            checklist_item["weight_evaluation"] = weight_eval
                            checklist_item["score"] = weight_eval["evaluation"]["final_score"]

                    parent["checklist"].append(checklist_item)
                    i += 1
                else:
                    # 새 하위 토글 생성
                    child_title = self._clean_title(text)
                    current_child = {
                        "title": child_title[:100],  # 제목 길이 제한
                        "content": "",
                        "current_score": 0,
                        "max_score": 100,
                        "children": [],
                        "checklist": []
                    }
                    parent["children"].append(current_child)
                    i += 1

            # 더 하위 레벨인 경우 - 재귀 호출
            elif level > parent_level + 1:
                if current_child:
                    i = self._build_hierarchy(current_child, items, i, level - 1)
                else:
                    # 부모가 없으면 본문에 추가
                    if parent["content"]:
                        parent["content"] += "\n\n"
                    parent["content"] += text
                    i += 1
            else:
                i += 1

        return i

    def _clean_title(self, text: str) -> str:
        """제목 텍스트 정리"""
        # 번호 패턴 제거
        text = re.sub(r'^[IVX]+\.\s+', '', text)
        text = re.sub(r'^\d+\.\s+', '', text)
        text = re.sub(r'^\d+\.\d+\.\s+', '', text)
        text = re.sub(r'^\d+\.\d+\.\d+\.\s+', '', text)
        text = re.sub(r'^\(\d+\)\s+', '', text)
        text = re.sub(r'^[가-힣]\)\s+', '', text)

        return text.strip()

    def _clean_list_text(self, text: str) -> str:
        """목록 텍스트 정리"""
        # 목록 기호 제거
        text = re.sub(r'^[-•·]\s+', '', text)
        text = re.sub(r'^\d+[.)]\s+', '', text)
        text = re.sub(r'^[가-힣][.)]\s+', '', text)

        return text.strip()

    def _auto_evaluate_checklist(self, text: str) -> Dict:
        """
        체크리스트 항목을 자동으로 평가

        Args:
            text: 체크리스트 항목 텍스트

        Returns:
            Dict: 가중치 평가 정보
        """
        if not self.evaluator:
            return None

        # 키워드 기반 자동 점수 부여
        scores = self._analyze_text_for_scores(text)

        # 평가 수행
        try:
            evaluation = self.evaluator.evaluate_checklist_item(
                c1_score=scores["C1"],
                c1_rationale=scores["C1_rationale"],
                c2_score=scores["C2"],
                c2_rationale=scores["C2_rationale"],
                c3_score=scores["C3"],
                c3_rationale=scores["C3_rationale"],
                c4_score=scores["C4"],
                c4_rationale=scores["C4_rationale"],
                c5_score=scores["C5"],
                c5_rationale=scores["C5_rationale"],
                uncertainty_factor=scores["U"],
                dependency_factor=scores["D"],
                regulatory_gate_flag=scores["G"]
            )

            result = self.evaluator.create_checklist_item_result(
                item_id=0,
                category=scores["category"],
                item=text,
                evaluation=evaluation
            )

            return result

        except Exception as e:
            print(f"가중치 평가 실패: {e}")
            return None

    def _analyze_text_for_scores(self, text: str) -> Dict:
        """
        텍스트 분석하여 자동으로 점수 산정

        Args:
            text: 분석할 텍스트

        Returns:
            Dict: 점수 및 근거
        """
        text_lower = text.lower()

        # 기본 점수
        scores = {
            "C1": 3, "C1_rationale": "일반적인 검토 항목",
            "C2": 3, "C2_rationale": "일반적인 비용/일정 영향",
            "C3": 3, "C3_rationale": "일반적인 환경/안전 고려사항",
            "C4": 3, "C4_rationale": "일반적인 운영 영향",
            "C5": 3, "C5_rationale": "일반적인 수정 난이도",
            "U": 1.0,
            "D": 1.0,
            "G": 0.0,
            "category": "일반"
        }

        # C1: 승인/법규 관문성
        approval_keywords = ["승인", "인허가", "허가", "면허", "등록", "신고", "협의", "법정", "규제"]
        if any(k in text for k in approval_keywords):
            scores["C1"] = 4
            scores["C1_rationale"] = "인허가 또는 승인 관련 항목"
            scores["category"] = "승인/규제"
            if "필수" in text or "법정" in text:
                scores["C1"] = 5
                scores["C1_rationale"] = "법정 필수 승인 항목"
                scores["G"] = 0.5

        # C2: 비용/일정 영향
        cost_keywords = ["비용", "예산", "capex", "opex", "투자", "지출"]
        schedule_keywords = ["일정", "공정", "지연", "납기", "완료", "기한"]
        if any(k in text_lower for k in cost_keywords):
            scores["C2"] = 4
            scores["C2_rationale"] = "비용 영향이 있는 항목"
            scores["category"] = "비용"
        if any(k in text_lower for k in schedule_keywords):
            scores["C2"] = max(scores["C2"], 4)
            scores["C2_rationale"] = "일정 영향이 있는 항목"

        # C3: 환경·안전 영향
        env_keywords = ["환경", "eia", "환경영향평가", "소음", "대기", "수질", "폐기물", "민원"]
        safety_keywords = ["안전", "위험", "사고", "재해", "보안", "화재", "방재"]
        if any(k in text_lower for k in env_keywords):
            scores["C3"] = 4
            scores["C3_rationale"] = "환경 영향이 있는 항목"
            scores["category"] = "환경"
            if "환경영향평가" in text or "eia" in text_lower:
                scores["C3"] = 5
                scores["C3_rationale"] = "환경영향평가 관련 핵심 항목"
        if any(k in text for k in safety_keywords):
            scores["C3"] = max(scores["C3"], 4)
            scores["C3_rationale"] = "안전 관련 항목"

        # C4: 운영성 영향
        operation_keywords = ["운영", "otp", "수하물", "회전율", "용량", "처리량", "서비스", "효율"]
        if any(k in text_lower for k in operation_keywords):
            scores["C4"] = 4
            scores["C4_rationale"] = "운영에 영향을 미치는 항목"
            scores["category"] = "운영"
            if "용량" in text or "처리량" in text:
                scores["C4"] = 5
                scores["C4_rationale"] = "공항 용량에 치명적 영향"

        # C5: 대체/가역성
        irreversible_keywords = ["건설", "구조물", "인프라", "설계", "배치", "레이아웃", "설치"]
        if any(k in text for k in irreversible_keywords):
            scores["C5"] = 4
            scores["C5_rationale"] = "구조적 변경으로 수정이 어려움"
            if "건설" in text or "구조물" in text:
                scores["C5"] = 5
                scores["C5_rationale"] = "건설 후 수정 불가능"

        # 불확실성 계수
        if "계획" in text or "검토" in text:
            scores["U"] = 1.1  # 아직 확정되지 않아 불확실성 있음

        # 의존성 계수
        if "기본" in text or "핵심" in text or "주요" in text:
            scores["D"] = 1.2  # 다른 결정에 영향을 미치는 허브성

        return scores

    def process_excel(self, excel_path: str) -> Dict:
        """
        Excel 파일을 처리하여 토글 구조로 변환

        Args:
            excel_path: Excel 파일 경로

        Returns:
            Dict: 토글 구조 데이터
        """
        filename = os.path.basename(excel_path)
        filename_without_ext = os.path.splitext(filename)[0]

        # 1. 데이터 추출
        rows_data = self.extract_data_from_excel(excel_path)

        if not rows_data:
            return None

        # 템플릿 사용 시
        if self.use_template and self.template_manager:
            # 전체 데이터를 문자열로 합치기
            content_lines = []
            for row in rows_data:
                row_text = " | ".join([str(cell) for cell in row if cell])
                if row_text.strip():
                    content_lines.append(row_text)

            content = "\n".join(content_lines)

            # 템플릿 적용
            toggle_data = self.template_manager.create_project_from_file(
                filename_without_ext,
                content
            )

            return toggle_data

        # 템플릿 미사용 시 (기존 방식)
        # 2. 구조 분석
        structured_items = self.detect_structure(rows_data)

        # 3. 토글 구조로 변환
        toggle_data = self.convert_to_toggle_structure(structured_items)

        return toggle_data


def is_excel_supported() -> bool:
    """Excel 처리가 지원되는지 확인"""
    return EXCEL_SUPPORT
