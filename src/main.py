import sys
import os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
os.environ['QT_QPA_PLATFORM_PLUGIN_PATH'] = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'venv', 'Lib', 'site-packages', 'PyQt5', 'Qt5', 'plugins')
import json
from typing import Dict, List
from PyQt5.QtWidgets import (QApplication, QMainWindow, QVBoxLayout, QHBoxLayout,
                             QWidget, QScrollArea, QMenuBar, QAction, QFileDialog,
                             QMessageBox, QStatusBar, QPushButton, QFrame, QLabel)
from PyQt5.QtCore import Qt, QTimer, pyqtSignal, QUrl, QMimeData
from PyQt5.QtGui import QPalette, QColor, QDrag, QPainter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from models.toggle_item import ToggleItem
# from components.toggle_widget import ToggleWidget  # 구 노션 스타일 (복잡함)
from components.simple_toggle_widget import SimpleToggleWidget as ToggleWidget  # 새 심플 스타일
from commands.base_command import CommandHistory


class RiskManagementApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.current_file = None
        self.root_items: List[ToggleItem] = []
        self.root_widgets: List[ToggleWidget] = []
        self.is_modified = False
        self.separator_widget = None
        self.loaded_items_count = 0
        self.selected_widget = None
        self.locked_files = {}  # 파일 경로 -> 파일 핸들 매핑

        # Command History for Undo/Redo
        self.command_history = CommandHistory()

        # Training Data Collector
        try:
            from utils.training_data_collector import TrainingDataCollector
            self.training_collector = TrainingDataCollector()
        except ImportError:
            self.training_collector = None

        self.setup_ui()
        self.setup_menu()
        self.setup_auto_save()
        self.setup_dark_theme()
    
    def setup_ui(self):
        self.setWindowTitle("리스크 관리 시스템")
        self.setGeometry(100, 100, 1000, 700)
        
        # 드래그 앤 드롭 활성화
        self.setAcceptDrops(True)
        
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        main_layout = QVBoxLayout(central_widget)
        main_layout.setContentsMargins(10, 10, 10, 10)
        
        toolbar_layout = QHBoxLayout()

        # Undo/Redo 버튼
        self.undo_btn = QPushButton("⟲ 실행취소")
        self.undo_btn.setToolTip("마지막 작업을 취소합니다 (Ctrl+Z)")
        self.undo_btn.clicked.connect(self.undo_action)
        self.undo_btn.setEnabled(False)
        toolbar_layout.addWidget(self.undo_btn)

        self.redo_btn = QPushButton("⟳ 다시실행")
        self.redo_btn.setToolTip("취소된 작업을 다시 실행합니다 (Ctrl+Y)")
        self.redo_btn.clicked.connect(self.redo_action)
        self.redo_btn.setEnabled(False)
        toolbar_layout.addWidget(self.redo_btn)

        # 구분선
        separator = QFrame()
        separator.setFrameShape(QFrame.VLine)
        separator.setFrameShadow(QFrame.Sunken)
        toolbar_layout.addWidget(separator)

        add_root_btn = QPushButton("새 루트 토글 추가")
        add_root_btn.clicked.connect(self.add_root_toggle)
        toolbar_layout.addWidget(add_root_btn)

        # 구분선
        separator2 = QFrame()
        separator2.setFrameShape(QFrame.VLine)
        separator2.setFrameShadow(QFrame.Sunken)
        toolbar_layout.addWidget(separator2)

        # AI 분석 버튼
        progress_btn = QPushButton("📊 진행 상황 분석")
        progress_btn.setToolTip("AI가 전체 프로젝트 진행 상황을 분석합니다")
        progress_btn.clicked.connect(self.analyze_progress)
        toolbar_layout.addWidget(progress_btn)

        priority_btn = QPushButton("🎯 우선순위 정렬")
        priority_btn.setToolTip("AI가 작업 우선순위를 분석하여 자동으로 정렬합니다")
        priority_btn.clicked.connect(self.analyze_and_sort_priority)
        toolbar_layout.addWidget(priority_btn)

        toolbar_layout.addStretch()
        main_layout.addLayout(toolbar_layout)
        
        self.scroll_area = QScrollArea()
        self.scroll_area.setWidgetResizable(True)
        self.scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        
        self.scroll_widget = QWidget()
        self.scroll_layout = QVBoxLayout(self.scroll_widget)
        self.scroll_layout.setAlignment(Qt.AlignTop)
        self.scroll_layout.setSpacing(5)
        
        self.scroll_area.setWidget(self.scroll_widget)
        main_layout.addWidget(self.scroll_area)
        
        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)
        self.status_bar.showMessage("준비됨")

    def update_undo_redo_state(self):
        """Undo/Redo 버튼과 메뉴 상태 업데이트"""
        can_undo = self.command_history.can_undo()
        can_redo = self.command_history.can_redo()

        self.undo_btn.setEnabled(can_undo)
        self.redo_btn.setEnabled(can_redo)
        self.undo_menu_action.setEnabled(can_undo)
        self.redo_menu_action.setEnabled(can_redo)

        # 툴팁 업데이트
        if can_undo:
            undo_desc = self.command_history.get_undo_description()
            self.undo_btn.setToolTip(f"실행취소: {undo_desc} (Ctrl+Z)")
            self.undo_menu_action.setText(f"실행취소: {undo_desc}")
        else:
            self.undo_btn.setToolTip("실행취소 (Ctrl+Z)")
            self.undo_menu_action.setText("실행취소")

        if can_redo:
            redo_desc = self.command_history.get_redo_description()
            self.redo_btn.setToolTip(f"다시실행: {redo_desc} (Ctrl+Y)")
            self.redo_menu_action.setText(f"다시실행: {redo_desc}")
        else:
            self.redo_btn.setToolTip("다시실행 (Ctrl+Y)")
            self.redo_menu_action.setText("다시실행")

    def undo_action(self):
        """Undo 실행"""
        if self.command_history.undo():
            self.update_undo_redo_state()
            self.mark_modified()

    def redo_action(self):
        """Redo 실행"""
        if self.command_history.redo():
            self.update_undo_redo_state()
            self.mark_modified()

    def execute_command(self, command):
        """명령 실행 및 히스토리에 추가"""
        self.command_history.execute_command(command)
        self.update_undo_redo_state()

    def setup_menu(self):
        menubar = self.menuBar()

        # 편집 메뉴
        edit_menu = menubar.addMenu('편집')

        self.undo_menu_action = QAction('실행취소', self)
        self.undo_menu_action.setShortcut('Ctrl+Z')
        self.undo_menu_action.triggered.connect(self.undo_action)
        self.undo_menu_action.setEnabled(False)
        edit_menu.addAction(self.undo_menu_action)

        self.redo_menu_action = QAction('다시실행', self)
        self.redo_menu_action.setShortcut('Ctrl+Y')
        self.redo_menu_action.triggered.connect(self.redo_action)
        self.redo_menu_action.setEnabled(False)
        edit_menu.addAction(self.redo_menu_action)

        edit_menu.addSeparator()

        file_menu = menubar.addMenu('파일')
        
        new_action = QAction('새 프로젝트', self)
        new_action.setShortcut('Ctrl+N')
        new_action.triggered.connect(self.new_project)
        file_menu.addAction(new_action)
        
        open_folder_action = QAction('폴더에서 토글 불러오기', self)
        open_folder_action.setShortcut('Ctrl+O')
        open_folder_action.triggered.connect(self.open_toggle_folder)
        file_menu.addAction(open_folder_action)
        
        save_all_action = QAction('모든 토글을 폴더에 저장', self)
        save_all_action.setShortcut('Ctrl+Shift+S')
        save_all_action.triggered.connect(self.save_all_toggles_to_folder)
        file_menu.addAction(save_all_action)
        
        file_menu.addSeparator()
        
        export_action = QAction('Excel로 내보내기', self)
        export_action.triggered.connect(self.export_to_excel)
        file_menu.addAction(export_action)

        # 프린트 메뉴
        print_preview_action = QAction('체크리스트 프린트 미리보기', self)
        print_preview_action.setShortcut('Ctrl+Shift+P')
        print_preview_action.triggered.connect(self.print_preview_checklist)
        file_menu.addAction(print_preview_action)

        print_action = QAction('체크리스트 프린트', self)
        print_action.setShortcut('Ctrl+P')
        print_action.triggered.connect(self.print_checklist)
        file_menu.addAction(print_action)

        export_pdf_action = QAction('체크리스트 PDF로 저장', self)
        export_pdf_action.triggered.connect(self.export_checklist_to_pdf)
        file_menu.addAction(export_pdf_action)

        file_menu.addSeparator()
        
        exit_action = QAction('종료', self)
        exit_action.setShortcut('Ctrl+Q')
        exit_action.triggered.connect(self.close)
        file_menu.addAction(exit_action)

        # AI 학습 메뉴
        if self.training_collector:
            ai_menu = menubar.addMenu('AI 학습')

            save_training_action = QAction('현재 작업을 학습 데이터로 저장', self)
            save_training_action.setShortcut('Ctrl+L')
            save_training_action.triggered.connect(self.save_current_as_training_data)
            ai_menu.addAction(save_training_action)

            ai_menu.addSeparator()

            stats_action = QAction('학습 데이터 통계 보기', self)
            stats_action.triggered.connect(self.show_training_statistics)
            ai_menu.addAction(stats_action)

            ai_menu.addSeparator()

            export_openai_action = QAction('내보내기: OpenAI 파인튜닝 형식', self)
            export_openai_action.triggered.connect(self.export_openai_format)
            ai_menu.addAction(export_openai_action)

            export_local_action = QAction('내보내기: 로컬 파인튜닝 형식', self)
            export_local_action.triggered.connect(self.export_local_format)
            ai_menu.addAction(export_local_action)

    def setup_auto_save(self):
        self.auto_save_timer = QTimer()
        self.auto_save_timer.timeout.connect(self.auto_save)
        self.auto_save_timer.start(60000)
    
    def setup_dark_theme(self):
        # Notion 라이트 모드 스타일
        self.setStyleSheet("""
            QMainWindow {
                background-color: #ffffff;
                color: #37352f;
            }
            QMenuBar {
                background-color: #fbfbfa;
                color: #37352f;
                border-bottom: 1px solid #e9e9e7;
                padding: 2px;
            }
            QMenuBar::item {
                background-color: transparent;
                padding: 6px 12px;
                border-radius: 3px;
            }
            QMenuBar::item:selected {
                background-color: #f1f1ef;
            }
            QMenu {
                background-color: #ffffff;
                color: #37352f;
                border: 1px solid #e9e9e7;
                border-radius: 4px;
                padding: 4px;
            }
            QMenu::item {
                padding: 6px 12px;
                border-radius: 3px;
            }
            QMenu::item:selected {
                background-color: #f1f1ef;
            }
            QScrollArea {
                background-color: #ffffff;
                border: none;
            }
            QStatusBar {
                background-color: #fbfbfa;
                color: #787774;
                border-top: 1px solid #e9e9e7;
            }
            QPushButton {
                background-color: rgba(55, 53, 47, 0.08);
                border: none;
                border-radius: 3px;
                color: #37352f;
                padding: 6px 12px;
                font-weight: 500;
            }
            QPushButton:hover {
                background-color: rgba(55, 53, 47, 0.12);
            }
            QPushButton:pressed {
                background-color: rgba(55, 53, 47, 0.16);
            }
            QPushButton:disabled {
                background-color: rgba(55, 53, 47, 0.04);
                color: rgba(55, 53, 47, 0.4);
            }
        """)
    
    def add_initial_toggle(self):
        root_item = ToggleItem("프로젝트 루트", "새 리스크 관리 프로젝트입니다.")
        self.root_items.append(root_item)
        self.add_widget_for_item(root_item, None)
    
    def add_root_toggle(self):
        from commands.toggle_commands import AddToggleCommand
        # 첫 번째 새 토글 추가 시 구분선 생성
        if len(self.root_items) >= self.loaded_items_count and not self.separator_widget:
            self.create_separator()

        command = AddToggleCommand(self, None, {"title": "새 루트 토글", "content": ""})
        self.execute_command(command)
    
    def add_widget_for_item(self, item: ToggleItem, parent_widget: ToggleWidget = None, include_children: bool = True):
        widget = ToggleWidget(item)
        widget.item_changed.connect(self.on_item_changed)
        widget.delete_requested.connect(self.on_delete_requested)
        widget.add_child_requested.connect(self.on_add_child_requested)
        widget.save_requested.connect(self.on_save_requested)
        widget.selection_changed.connect(self.on_selection_changed)
        widget.move_requested.connect(self.on_move_requested)
        
        if parent_widget is None:
            self.root_widgets.append(widget)
            self.scroll_layout.addWidget(widget)
        else:
            parent_widget.add_child_widget(widget)
        
        if include_children:
            for child_item in item.children:
                self.add_widget_for_item(child_item, widget, True)
        
        return widget
    
    def on_item_changed(self):
        self.mark_modified()
        self.update_status()
        # 모든 위젯의 진행률 표시를 업데이트 (부모 위젯 포함)
        self.update_all_widget_displays()

    def update_all_widget_displays(self):
        """모든 위젯의 진행률 표시를 업데이트 (부모 위젯도 포함)"""
        def update_widget_recursive(widget):
            widget.update_progress_display()
            for child_widget in widget.child_widgets:
                update_widget_recursive(child_widget)

        for root_widget in self.root_widgets:
            update_widget_recursive(root_widget)

    def on_delete_requested(self, widget: ToggleWidget):
        reply = QMessageBox.question(self, '확인', '정말로 이 토글을 삭제하시겠습니까?',
                                   QMessageBox.Yes | QMessageBox.No, QMessageBox.No)

        if reply == QMessageBox.Yes:
            from commands.toggle_commands import DeleteToggleCommand
            command = DeleteToggleCommand(self, widget)
            self.execute_command(command)
    
    def on_add_child_requested(self, parent_widget: ToggleWidget):
        from commands.toggle_commands import AddToggleCommand
        command = AddToggleCommand(self, parent_widget, {"title": "새 하위 토글", "content": ""})
        self.execute_command(command)
    
    def on_save_requested(self, widget: ToggleWidget):
        filename = None
        
        # 이미 저장된 파일이 있으면 그 파일로 저장
        if widget.item.source_file and os.path.exists(widget.item.source_file):
            filename = widget.item.source_file
        else:
            # 새 파일로 저장
            safe_title = "".join(c for c in widget.item.title if c.isalnum() or c in (' ', '-', '_')).rstrip()
            if not safe_title:
                safe_title = "토글"
            
            default_filename = f"{safe_title}.json"
            filename, _ = QFileDialog.getSaveFileName(self, '토글 저장', default_filename,
                                                     'JSON Files (*.json);;All Files (*)')
        
        if filename:
            try:
                # 개별 토글 데이터 생성
                data = {
                    'root_items': [widget.item.to_dict()],
                    'version': '1.0',
                    'is_single_toggle': True,
                    'original_title': widget.item.title
                }
                
                with open(filename, 'w', encoding='utf-8') as f:
                    json.dump(data, f, ensure_ascii=False, indent=2)
                
                # 저장된 파일명을 토글에 설정하고 UI 업데이트
                widget.item.source_file = filename
                widget.update_source_file_display()
                
                # 새로 저장된 파일 잠금
                self.lock_file(filename)
                
                if widget.item.source_file == filename:
                    self.status_bar.showMessage(f"토글 업데이트 완료: {os.path.basename(filename)}")
                else:
                    self.status_bar.showMessage(f"토글 저장 완료: {os.path.basename(filename)}")
                
            except Exception as e:
                QMessageBox.critical(self, '저장 오류', f'파일을 저장할 수 없습니다:\n{str(e)}')
    
    def on_selection_changed(self, widget: ToggleWidget):
        # 이전에 선택된 위젯의 선택 해제
        if self.selected_widget and self.selected_widget != widget:
            self.selected_widget.set_selected(False)
        
        # 새로운 위젯 선택
        self.selected_widget = widget
    
    def on_move_requested(self, target_widget: ToggleWidget, dragged_widget_id: int):
        """토글 위젯 순서 변경 요청 처리"""
        # 드래그된 위젯 찾기
        dragged_widget = None
        for widget in self.root_widgets:
            if id(widget) == dragged_widget_id:
                dragged_widget = widget
                break
        
        if not dragged_widget or dragged_widget == target_widget:
            return
        
        # 현재 위치와 새 위치 찾기
        old_index = self.root_widgets.index(dragged_widget)
        new_index = self.root_widgets.index(target_widget)
        
        # 리스트에서 위젯과 아이템 이동
        dragged_item = self.root_items.pop(old_index)
        self.root_widgets.pop(old_index)
        
        # 레이아웃에서 위젯 제거
        self.scroll_layout.removeWidget(dragged_widget)
        
        # 새 위치에 삽입
        self.root_items.insert(new_index, dragged_item)
        self.root_widgets.insert(new_index, dragged_widget)
        self.scroll_layout.insertWidget(new_index, dragged_widget)
        
        # 변경사항 표시
        self.mark_modified()
        self.status_bar.showMessage(f"토글 순서가 변경되었습니다: {old_index + 1} → {new_index + 1}")
        
        print(f"토글 이동: '{dragged_widget.item.title}' {old_index} → {new_index}")
    
    def delete_widget(self, widget: ToggleWidget):
        if widget in self.root_widgets:
            self.root_widgets.remove(widget)
            self.root_items.remove(widget.item)
            self.scroll_layout.removeWidget(widget)
        else:
            parent_widget = self.find_parent_widget(widget)
            if parent_widget:
                parent_widget.remove_child_widget(widget)
                parent_widget.item.remove_child(widget.item)
        
        widget.setParent(None)
        self.mark_modified()
    
    def lock_file(self, file_path: str):
        """파일을 잠가서 외부에서 삭제할 수 없도록 함"""
        try:
            if file_path and file_path not in self.locked_files:
                # 파일을 읽기+쓰기 모드로 열어서 잠금
                file_handle = open(file_path, 'r+', encoding='utf-8')
                self.locked_files[file_path] = file_handle
                print(f"파일 잠금: {file_path}")
        except Exception as e:
            print(f"파일 잠금 실패: {file_path}, 오류: {e}")
    
    def unlock_file(self, file_path: str):
        """파일 잠금 해제"""
        try:
            if file_path in self.locked_files:
                self.locked_files[file_path].close()
                del self.locked_files[file_path]
                print(f"파일 잠금 해제: {file_path}")
        except Exception as e:
            print(f"파일 잠금 해제 실패: {file_path}, 오류: {e}")
    
    def unlock_all_files(self):
        """모든 파일 잠금 해제"""
        for file_path in list(self.locked_files.keys()):
            self.unlock_file(file_path)
    
    def find_parent_widget(self, target_widget: ToggleWidget) -> ToggleWidget:
        for root_widget in self.root_widgets:
            parent = self._find_parent_widget_recursive(root_widget, target_widget)
            if parent:
                return parent
        return None
    
    def _find_parent_widget_recursive(self, current_widget: ToggleWidget, 
                                    target_widget: ToggleWidget) -> ToggleWidget:
        if target_widget in current_widget.child_widgets:
            return current_widget
        
        for child_widget in current_widget.child_widgets:
            parent = self._find_parent_widget_recursive(child_widget, target_widget)
            if parent:
                return parent
        
        return None
    
    def mark_modified(self):
        self.is_modified = True
        title = "리스크 관리 시스템"
        if self.current_file:
            title += f" - {os.path.basename(self.current_file)}"
        title += " *"
        self.setWindowTitle(title)
    
    def update_status(self):
        total_items = len(self.root_items)
        for root_item in self.root_items:
            total_items += len(root_item.get_all_descendants())
        
        total_score = sum(item.get_total_score() for item in self.root_items)
        total_max = sum(item.get_total_max_score() for item in self.root_items)
        
        self.status_bar.showMessage(f"토글: {total_items}개 | 전체 진행률: {total_score}/{total_max}")
    
    def create_separator(self):
        if self.separator_widget:
            return
            
        separator_container = QWidget()
        separator_layout = QVBoxLayout(separator_container)
        separator_layout.setContentsMargins(20, 10, 20, 10)
        
        # 구분선
        line = QFrame()
        line.setFrameShape(QFrame.HLine)
        line.setFrameShadow(QFrame.Sunken)
        line.setStyleSheet("QFrame { color: #555; }")
        
        # 레이블
        label = QLabel("--- 새로 추가된 토글 ---")
        label.setAlignment(Qt.AlignCenter)
        label.setStyleSheet("QLabel { color: #888; font-style: italic; margin: 5px; }")
        
        separator_layout.addWidget(line)
        separator_layout.addWidget(label)
        
        self.separator_widget = separator_container
        self.scroll_layout.addWidget(self.separator_widget)
    
    def remove_separator(self):
        if self.separator_widget:
            self.scroll_layout.removeWidget(self.separator_widget)
            self.separator_widget.setParent(None)
            self.separator_widget = None
    
    def new_project(self):
        if self.is_modified:
            reply = QMessageBox.question(self, '새 프로젝트', 
                                       '변경사항이 있습니다. 저장하지 않고 계속하시겠습니까?',
                                       QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            if reply != QMessageBox.Yes:
                return
        
        self.clear_project()
        self.is_modified = False
        self.setWindowTitle("리스크 관리 시스템")
    
    def clear_project(self):
        # 모든 파일 잠금 해제
        self.unlock_all_files()
        
        for widget in self.root_widgets[:]:
            widget.setParent(None)
        self.root_widgets.clear()
        self.root_items.clear()
        self.remove_separator()
        self.loaded_items_count = 0
    
    def open_toggle_folder(self):
        """폴더에서 JSON 토글 파일들을 불러오기"""
        folder = QFileDialog.getExistingDirectory(self, '토글 파일들이 있는 폴더 선택')
        if not folder:
            return
        
        # 기존 토글들 제거
        self.clear_project()
        
        try:
            json_files = []
            for file_name in os.listdir(folder):
                if file_name.lower().endswith('.json'):
                    json_files.append(os.path.join(folder, file_name))
            
            if not json_files:
                QMessageBox.information(self, '파일 없음', '선택한 폴더에 JSON 파일이 없습니다.')
                return
            
            loaded_count = 0
            errors = []
            
            for file_path in json_files:
                try:
                    with open(file_path, 'r', encoding='utf-8') as f:
                        data = json.load(f)
                    
                    for item_data in data.get('root_items', []):
                        root_item = ToggleItem.from_dict(item_data)
                        root_item.source_file = file_path
                        self.root_items.append(root_item)
                        self.add_widget_for_item(root_item, None)
                        loaded_count += 1
                    
                    # 파일 잠금
                    self.lock_file(file_path)
                    
                except Exception as e:
                    errors.append(f"{os.path.basename(file_path)}: {str(e)}")
            
            if loaded_count > 0:
                self.loaded_items_count = len(self.root_items)
                self.setWindowTitle(f"리스크 관리 시스템 - {os.path.basename(folder)}")
                self.status_bar.showMessage(f'{loaded_count}개 토글을 폴더에서 불러왔습니다.')
                self.update_status()
                
                if errors:
                    QMessageBox.warning(self, '일부 파일 로드 실패', 
                                      f'{loaded_count}개 토글이 로드되었지만 일부 파일에서 오류가 발생했습니다:\n\n' + 
                                      '\n'.join(errors))
            elif errors:
                QMessageBox.critical(self, '폴더 로드 실패', 
                                   '파일을 불러올 수 없습니다:\n\n' + '\n'.join(errors))
        
        except Exception as e:
            QMessageBox.critical(self, '폴더 로드 오류', f'폴더를 읽을 수 없습니다:\n{str(e)}')
    
    def save_all_toggles_to_folder(self):
        """모든 토글을 개별 파일로 폴더에 저장"""
        if not self.root_items:
            QMessageBox.information(self, '저장 불가', '저장할 토글이 없습니다.')
            return
        
        folder = QFileDialog.getExistingDirectory(self, '토글 파일들을 저장할 폴더 선택')
        if not folder:
            return
        
        try:
            saved_files = []
            updated_files = []
            new_files = []
            
            for i, root_item in enumerate(self.root_items):
                filename = None
                
                # 이미 저장된 파일이 있는지 확인
                if root_item.source_file and os.path.exists(root_item.source_file):
                    # 기존 파일이 선택된 폴더에 있으면 업데이트
                    if os.path.dirname(root_item.source_file) == folder:
                        filename = root_item.source_file
                        updated_files.append(os.path.basename(filename))
                    else:
                        # 다른 폴더의 파일이면 새 파일명으로 저장
                        existing_name = os.path.basename(root_item.source_file)
                        filename = os.path.join(folder, existing_name)
                        
                        # 같은 이름 파일이 있으면 번호 추가
                        counter = 1
                        original_filename = filename
                        while os.path.exists(filename):
                            name_without_ext = os.path.splitext(original_filename)[0]
                            filename = f"{name_without_ext}_{counter}.json"
                            counter += 1
                        new_files.append(os.path.basename(filename))
                else:
                    # 새 파일로 저장
                    safe_title = "".join(c for c in root_item.title if c.isalnum() or c in (' ', '-', '_')).rstrip()
                    if not safe_title:
                        safe_title = f"토글_{i+1}"
                    
                    filename = os.path.join(folder, f"{safe_title}.json")
                    
                    # 같은 이름 파일이 있으면 번호 추가
                    counter = 1
                    original_filename = filename
                    while os.path.exists(filename):
                        name_without_ext = os.path.splitext(original_filename)[0]
                        filename = f"{name_without_ext}_{counter}.json"
                        counter += 1
                    new_files.append(os.path.basename(filename))
                
                # 토글 데이터 저장
                data = {
                    'root_items': [root_item.to_dict()],
                    'version': '1.0',
                    'is_single_toggle': True,
                    'original_title': root_item.title
                }
                
                with open(filename, 'w', encoding='utf-8') as f:
                    json.dump(data, f, ensure_ascii=False, indent=2)
                
                # 토글에 파일 경로 연결하고 잠금
                old_file = root_item.source_file
                root_item.source_file = filename
                
                # 이전 파일 잠금 해제하고 새 파일 잠금
                if old_file and old_file != filename:
                    self.unlock_file(old_file)
                self.lock_file(filename)
                
                saved_files.append(os.path.basename(filename))
            
            # UI 업데이트
            for widget in self.root_widgets:
                widget.update_source_file_display()
            
            # 결과 메시지 생성
            message = f'{len(saved_files)}개의 토글이 저장되었습니다:\n\n'
            if updated_files:
                message += f'업데이트된 파일 ({len(updated_files)}개):\n' + '\n'.join(updated_files) + '\n\n'
            if new_files:
                message += f'새로 생성된 파일 ({len(new_files)}개):\n' + '\n'.join(new_files)
            
            QMessageBox.information(self, '저장 완료', message)
            
        except Exception as e:
            QMessageBox.critical(self, '저장 오류', f'파일을 저장할 수 없습니다:\n{str(e)}')
    
    def auto_save(self):
        """변경된 토글들을 자동으로 개별 파일에 저장"""
        if not self.is_modified:
            return
        
        auto_saved_count = 0
        for root_item in self.root_items:
            if root_item.source_file:
                try:
                    data = {
                        'root_items': [root_item.to_dict()],
                        'version': '1.0',
                        'is_single_toggle': True,
                        'original_title': root_item.title
                    }
                    
                    with open(root_item.source_file, 'w', encoding='utf-8') as f:
                        json.dump(data, f, ensure_ascii=False, indent=2)
                    
                    auto_saved_count += 1
                    
                except Exception as e:
                    print(f"자동 저장 실패: {root_item.source_file}, 오류: {e}")
        
        if auto_saved_count > 0:
            self.status_bar.showMessage(f"자동 저장 완료: {auto_saved_count}개 토글", 2000)
    
    def export_to_excel(self):
        if not self.root_items:
            QMessageBox.warning(self, '내보내기 오류', '내보낼 데이터가 없습니다.')
            return
        
        filename, _ = QFileDialog.getSaveFileName(self, 'Excel로 내보내기', '', 
                                                 'Excel Files (*.xlsx);;All Files (*)')
        if filename:
            self._export_to_excel(filename)
    
    def _export_to_excel(self, filename: str):
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "리스크 관리"
            
            headers = ['레벨', '제목', '현재 점수', '최대 점수', '진행률', '상세 내용']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')
            
            row = 2
            for root_item in self.root_items:
                row = self._write_item_to_excel(ws, root_item, 0, row)
            
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(filename)
            self.status_bar.showMessage(f"Excel 내보내기 완료: {filename}")
            QMessageBox.information(self, '내보내기 완료', f'Excel 파일로 성공적으로 내보냈습니다:\n{filename}')
        
        except Exception as e:
            QMessageBox.critical(self, '내보내기 오류', f'Excel 파일을 만들 수 없습니다:\n{str(e)}')
    
    def _write_item_to_excel(self, ws, item: ToggleItem, level: int, row: int) -> int:
        indent = "  " * level
        ws.cell(row=row, column=1, value=level)
        ws.cell(row=row, column=2, value=f"{indent}{item.title}")
        ws.cell(row=row, column=3, value=item.get_total_score())
        ws.cell(row=row, column=4, value=item.get_total_max_score())
        
        percentage = item.get_completion_percentage()
        ws.cell(row=row, column=5, value=f"{percentage:.1f}%")
        ws.cell(row=row, column=6, value=item.content)
        
        row += 1
        for child in item.children:
            row = self._write_item_to_excel(ws, child, level + 1, row)
        
        return row
    
    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            urls = event.mimeData().urls()
            # JSON, PDF, Word, Excel 파일인지 확인
            if any(url.toLocalFile().lower().endswith(('.json', '.pdf', '.docx', '.doc', '.xlsx', '.xls')) for url in urls):
                event.acceptProposedAction()
            else:
                event.ignore()
        else:
            event.ignore()
    
    def dragMoveEvent(self, event):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
        else:
            event.ignore()
    
    def dropEvent(self, event):
        if event.mimeData().hasUrls():
            urls = event.mimeData().urls()
            json_files = [url.toLocalFile() for url in urls if url.toLocalFile().lower().endswith('.json')]
            pdf_files = [url.toLocalFile() for url in urls if url.toLocalFile().lower().endswith('.pdf')]
            word_files = [url.toLocalFile() for url in urls if url.toLocalFile().lower().endswith(('.docx', '.doc'))]
            excel_files = [url.toLocalFile() for url in urls if url.toLocalFile().lower().endswith(('.xlsx', '.xls'))]

            if json_files:
                self.load_dropped_files(json_files)
                event.acceptProposedAction()
            elif pdf_files:
                self.load_dropped_pdf_files(pdf_files)
                event.acceptProposedAction()
            elif word_files:
                self.load_dropped_word_files(word_files)
                event.acceptProposedAction()
            elif excel_files:
                self.load_dropped_excel_files(excel_files)
                event.acceptProposedAction()
            else:
                event.ignore()
        else:
            event.ignore()
    
    def load_dropped_files(self, file_paths):
        if not file_paths:
            return
        
        loaded_count = 0
        errors = []
        
        for file_path in file_paths:
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                
                # 첫 번째 새 토글 추가 시 구분선 생성
                if len(self.root_items) >= self.loaded_items_count and not self.separator_widget:
                    self.create_separator()
                
                for item_data in data.get('root_items', []):
                    root_item = ToggleItem.from_dict(item_data)
                    root_item.source_file = file_path
                    self.root_items.append(root_item)
                    self.add_widget_for_item(root_item, None)
                    loaded_count += 1
                
                # 드래그 앤 드롭으로 로드한 파일도 잠금
                self.lock_file(file_path)
                
            except Exception as e:
                errors.append(f"{os.path.basename(file_path)}: {str(e)}")
        
        if loaded_count > 0:
            self.mark_modified()
            self.update_status()
            
            if errors:
                QMessageBox.warning(self, '일부 파일 로드 실패', 
                                  f'{loaded_count}개 토글이 추가되었지만 일부 파일에서 오류가 발생했습니다:\n\n' + 
                                  '\n'.join(errors))
            else:
                self.status_bar.showMessage(f'{loaded_count}개 토글이 드래그 앤 드롭으로 추가되었습니다.')
        elif errors:
            QMessageBox.critical(self, '파일 로드 실패', 
                               '파일을 불러올 수 없습니다:\n\n' + '\n'.join(errors))
    
    def load_dropped_pdf_files(self, file_paths):
        """PDF 파일을 드롭했을 때 토글로 변환"""
        if not file_paths:
            return

        try:
            from utils.pdf_processor import PDFProcessor, is_pdf_supported
            from components.page_range_dialog import PageRangeDialog

            if not is_pdf_supported():
                QMessageBox.critical(self, 'PDF 지원 불가',
                                   'PDF 처리를 위해 필요한 라이브러리가 설치되지 않았습니다.\n\n'
                                   'pip install pymupdf pdfplumber pillow')
                return
        except ImportError:
            QMessageBox.critical(self, 'PDF 지원 불가',
                               'PDF 처리를 위해 필요한 라이브러리가 설치되지 않았습니다.\n\n'
                               'pip install pymupdf pdfplumber pillow')
            return

        loaded_count = 0
        errors = []

        # 환경변수에서 LLM 모드 확인
        llm_mode = os.getenv("LLM_MODE", "none").lower()
        processor = PDFProcessor(llm_mode=llm_mode, use_template=True)

        for file_path in file_paths:
            try:
                # 페이지 수 확인
                total_pages = processor.get_page_count(file_path)
                if total_pages == 0:
                    errors.append(f"{os.path.basename(file_path)}: PDF를 열 수 없습니다")
                    continue

                # 페이지 범위 선택 다이얼로그 표시
                dialog_result = PageRangeDialog.get_page_range_from_user(
                    os.path.basename(file_path),
                    total_pages,
                    self
                )

                # 사용자가 취소한 경우
                if dialog_result is False:
                    continue

                # 페이지 범위와 검색어 추출
                page_range = dialog_result['page_range']
                search_keyword = dialog_result['search_keyword']

                self.status_bar.showMessage(f'PDF 처리 중: {os.path.basename(file_path)}...')

                # PDF 처리 (페이지 범위와 검색어 전달)
                toggle_data, error_msg = processor.process_pdf(file_path, page_range, search_keyword, self)

                if not toggle_data:
                    errors.append(f"{os.path.basename(file_path)}: {error_msg}")
                    continue

                # 첫 번째 새 토글 추가 시 구분선 생성
                if len(self.root_items) >= self.loaded_items_count and not self.separator_widget:
                    self.create_separator()

                # 토글 아이템 생성
                root_item = ToggleItem.from_dict(toggle_data)
                root_item.source_file = None  # PDF는 원본 파일로 저장하지 않음
                self.root_items.append(root_item)
                self.add_widget_for_item(root_item, None)
                loaded_count += 1

            except Exception as e:
                errors.append(f"{os.path.basename(file_path)}: {str(e)}")

        if loaded_count > 0:
            self.mark_modified()
            self.update_status()

            if errors:
                QMessageBox.warning(self, '일부 파일 변환 실패',
                                  f'{loaded_count}개 PDF가 토글로 변환되었지만 일부 파일에서 오류가 발생했습니다:\n\n' +
                                  '\n'.join(errors))
            else:
                self.status_bar.showMessage(f'{loaded_count}개 PDF가 토글로 변환되었습니다.')
        elif errors:
            QMessageBox.critical(self, 'PDF 변환 실패',
                               'PDF 파일을 변환할 수 없습니다:\n\n' + '\n'.join(errors))

    def load_dropped_word_files(self, file_paths):
        """Word 파일을 드롭했을 때 토글로 변환"""
        if not file_paths:
            return

        try:
            from utils.docx_processor import DOCXProcessor, is_docx_supported

            if not is_docx_supported():
                QMessageBox.critical(self, 'Word 지원 불가',
                                   'Word 처리를 위해 필요한 라이브러리가 설치되지 않았습니다.\n\n'
                                   'pip install python-docx')
                return
        except ImportError:
            QMessageBox.critical(self, 'Word 지원 불가',
                               'Word 처리를 위해 필요한 라이브러리가 설치되지 않았습니다.\n\n'
                               'pip install python-docx')
            return

        loaded_count = 0
        errors = []

        # 환경변수에서 LLM 모드 확인
        llm_mode = os.getenv("LLM_MODE", "none").lower()

        processor = DOCXProcessor(llm_mode=llm_mode, use_template=True)

        for file_path in file_paths:
            try:
                self.status_bar.showMessage(f'Word 처리 중: {os.path.basename(file_path)}...')

                # Word 처리
                toggle_data = processor.process_docx(file_path)

                if not toggle_data:
                    errors.append(f"{os.path.basename(file_path)}: Word에서 텍스트를 추출할 수 없습니다")
                    continue

                # 첫 번째 새 토글 추가 시 구분선 생성
                if len(self.root_items) >= self.loaded_items_count and not self.separator_widget:
                    self.create_separator()

                # 토글 아이템 생성
                root_item = ToggleItem.from_dict(toggle_data)
                root_item.source_file = None  # Word는 원본 파일로 저장하지 않음
                self.root_items.append(root_item)
                self.add_widget_for_item(root_item, None)
                loaded_count += 1

            except Exception as e:
                errors.append(f"{os.path.basename(file_path)}: {str(e)}")

        if loaded_count > 0:
            self.mark_modified()
            self.update_status()

            if errors:
                QMessageBox.warning(self, '일부 파일 변환 실패',
                                  f'{loaded_count}개 Word 파일이 토글로 변환되었지만 일부 파일에서 오류가 발생했습니다:\n\n' +
                                  '\n'.join(errors))
            else:
                self.status_bar.showMessage(f'{loaded_count}개 Word 파일이 토글로 변환되었습니다.')
        elif errors:
            QMessageBox.critical(self, 'Word 변환 실패',
                               'Word 파일을 변환할 수 없습니다:\n\n' + '\n'.join(errors))

    def load_dropped_excel_files(self, file_paths):
        """Excel 파일을 드롭했을 때 토글로 변환"""
        if not file_paths:
            return

        try:
            from utils.excel_processor import ExcelProcessor, is_excel_supported

            if not is_excel_supported():
                QMessageBox.critical(self, 'Excel 지원 불가',
                                   'Excel 처리를 위해 필요한 라이브러리가 설치되지 않았습니다.\n\n'
                                   'pip install openpyxl')
                return
        except ImportError:
            QMessageBox.critical(self, 'Excel 지원 불가',
                               'Excel 처리를 위해 필요한 라이브러리가 설치되지 않았습니다.\n\n'
                               'pip install openpyxl')
            return

        loaded_count = 0
        errors = []

        # 환경변수에서 LLM 모드 확인
        llm_mode = os.getenv("LLM_MODE", "none").lower()

        processor = ExcelProcessor(llm_mode=llm_mode, use_template=True)

        for file_path in file_paths:
            try:
                self.status_bar.showMessage(f'Excel 처리 중: {os.path.basename(file_path)}...')

                # Excel 처리
                toggle_data = processor.process_excel(file_path)

                if not toggle_data:
                    errors.append(f"{os.path.basename(file_path)}: Excel에서 데이터를 추출할 수 없습니다")
                    continue

                # 첫 번째 새 토글 추가 시 구분선 생성
                if len(self.root_items) >= self.loaded_items_count and not self.separator_widget:
                    self.create_separator()

                # 토글 아이템 생성
                root_item = ToggleItem.from_dict(toggle_data)
                root_item.source_file = None  # Excel은 원본 파일로 저장하지 않음
                self.root_items.append(root_item)
                self.add_widget_for_item(root_item, None)
                loaded_count += 1

            except Exception as e:
                errors.append(f"{os.path.basename(file_path)}: {str(e)}")

        if loaded_count > 0:
            self.mark_modified()
            self.update_status()

            if errors:
                QMessageBox.warning(self, '일부 파일 변환 실패',
                                  f'{loaded_count}개 Excel 파일이 토글로 변환되었지만 일부 파일에서 오류가 발생했습니다:\n\n' +
                                  '\n'.join(errors))
            else:
                self.status_bar.showMessage(f'{loaded_count}개 Excel 파일이 토글로 변환되었습니다.')
        elif errors:
            QMessageBox.critical(self, 'Excel 변환 실패',
                               'Excel 파일을 변환할 수 없습니다:\n\n' + '\n'.join(errors))

    def analyze_progress(self):
        """AI를 사용하여 진행 상황 분석"""
        if not self.root_items:
            QMessageBox.information(self, '분석 불가', '분석할 프로젝트가 없습니다.')
            return

        try:
            from utils.progress_analyzer import ProgressAnalyzer
            from components.analysis_dialog import AnalysisResultDialog, ProgressAnalysisDialog
            from PyQt5.QtCore import QThread, pyqtSignal

            # 분석 중 다이얼로그 표시
            progress_dialog = ProgressAnalysisDialog(self)

            # 백그라운드에서 분석 실행
            class AnalysisThread(QThread):
                finished = pyqtSignal(str)
                error = pyqtSignal(str)

                def __init__(self, root_items, llm_mode):
                    super().__init__()
                    self.root_items = root_items
                    self.llm_mode = llm_mode

                def run(self):
                    try:
                        analyzer = ProgressAnalyzer(llm_mode=self.llm_mode)
                        report = analyzer.generate_report(self.root_items)
                        self.finished.emit(report)
                    except Exception as e:
                        self.error.emit(str(e))

            # 환경변수에서 LLM 모드 가져오기
            llm_mode = os.getenv("LLM_MODE", "none").lower()

            self.analysis_thread = AnalysisThread(self.root_items, llm_mode)

            def on_analysis_finished(report):
                progress_dialog.close()
                result_dialog = AnalysisResultDialog("📊 프로젝트 진행 상황 분석", report, self)
                result_dialog.exec_()

            def on_analysis_error(error_msg):
                progress_dialog.close()
                QMessageBox.critical(self, '분석 오류', f'진행 상황 분석 중 오류가 발생했습니다:\n{error_msg}')

            self.analysis_thread.finished.connect(on_analysis_finished)
            self.analysis_thread.error.connect(on_analysis_error)
            self.analysis_thread.start()

            progress_dialog.exec_()

        except ImportError as e:
            QMessageBox.critical(self, '모듈 오류', f'필요한 모듈을 불러올 수 없습니다:\n{str(e)}')
        except Exception as e:
            QMessageBox.critical(self, '오류', f'진행 상황 분석 중 오류가 발생했습니다:\n{str(e)}')

    def analyze_and_sort_priority(self):
        """AI를 사용하여 우선순위 분석 및 정렬"""
        if not self.root_items:
            QMessageBox.information(self, '정렬 불가', '정렬할 프로젝트가 없습니다.')
            return

        try:
            from utils.priority_analyzer import PriorityAnalyzer
            from components.analysis_dialog import AnalysisResultDialog, ProgressAnalysisDialog
            from PyQt5.QtCore import QThread, pyqtSignal

            # 분석 중 다이얼로그 표시
            progress_dialog = ProgressAnalysisDialog(self)
            progress_dialog.setWindowTitle("우선순위 분석 중...")
            progress_dialog.update_status("AI가 작업 우선순위를 분석하고 있습니다...")

            # 백그라운드에서 분석 및 정렬 실행
            class PriorityThread(QThread):
                finished = pyqtSignal(list, str)
                error = pyqtSignal(str)

                def __init__(self, root_items, llm_mode):
                    super().__init__()
                    self.root_items = root_items
                    self.llm_mode = llm_mode

                def run(self):
                    try:
                        analyzer = PriorityAnalyzer(llm_mode=self.llm_mode)
                        sorted_items = analyzer.analyze_and_sort_items(self.root_items)
                        report = analyzer.get_priority_recommendations(self.root_items)
                        self.finished.emit(sorted_items, report)
                    except Exception as e:
                        self.error.emit(str(e))

            # 환경변수에서 LLM 모드 가져오기
            llm_mode = os.getenv("LLM_MODE", "none").lower()

            self.priority_thread = PriorityThread(self.root_items, llm_mode)

            def on_priority_finished(sorted_items, report):
                progress_dialog.close()

                # 사용자에게 정렬 여부 확인
                reply = QMessageBox.question(
                    self,
                    '우선순위 정렬',
                    f'AI가 {len(sorted_items)}개의 작업을 분석했습니다.\n\n'
                    '추천된 우선순위에 따라 작업을 재정렬하시겠습니까?\n'
                    '(현재 순서는 변경되지 않습니다)',
                    QMessageBox.Yes | QMessageBox.No,
                    QMessageBox.No
                )

                if reply == QMessageBox.Yes:
                    # 실제로 정렬 적용
                    self.apply_priority_sorting(sorted_items)
                    self.status_bar.showMessage("우선순위에 따라 작업이 재정렬되었습니다.")

                # 분석 결과 표시
                result_dialog = AnalysisResultDialog("🎯 우선순위 분석 보고서", report, self)
                result_dialog.exec_()

            def on_priority_error(error_msg):
                progress_dialog.close()
                QMessageBox.critical(self, '분석 오류', f'우선순위 분석 중 오류가 발생했습니다:\n{error_msg}')

            self.priority_thread.finished.connect(on_priority_finished)
            self.priority_thread.error.connect(on_priority_error)
            self.priority_thread.start()

            progress_dialog.exec_()

        except ImportError as e:
            QMessageBox.critical(self, '모듈 오류', f'필요한 모듈을 불러올 수 없습니다:\n{str(e)}')
        except Exception as e:
            QMessageBox.critical(self, '오류', f'우선순위 분석 중 오류가 발생했습니다:\n{str(e)}')

    def apply_priority_sorting(self, sorted_items):
        """우선순위 정렬을 실제로 적용"""
        # 기존 위젯들 제거
        for widget in self.root_widgets[:]:
            self.scroll_layout.removeWidget(widget)

        # 정렬된 순서대로 재배치
        self.root_items.clear()
        self.root_widgets.clear()

        for item, _priority_info in sorted_items:
            self.root_items.append(item)
            # 기존 위젯 찾기 또는 새로 생성
            widget = self.add_widget_for_item(item, None, include_children=False)
            self.root_widgets.append(widget)

        self.mark_modified()
        self.update_status()

    def print_preview_checklist(self):
        """체크리스트 프린트 미리보기"""
        if not self.root_items:
            QMessageBox.information(self, '프린트 불가', '프린트할 체크리스트가 없습니다.')
            return

        try:
            from utils.checklist_printer import ChecklistPrinter

            printer = ChecklistPrinter()
            printer.print_preview(self.root_items, self)
            self.status_bar.showMessage("프린트 미리보기를 표시했습니다.")

        except ImportError as e:
            QMessageBox.critical(self, '모듈 오류', f'프린트 모듈을 불러올 수 없습니다:\n{str(e)}')
        except Exception as e:
            QMessageBox.critical(self, '오류', f'프린트 미리보기 중 오류가 발생했습니다:\n{str(e)}')

    def print_checklist(self):
        """체크리스트 프린트"""
        if not self.root_items:
            QMessageBox.information(self, '프린트 불가', '프린트할 체크리스트가 없습니다.')
            return

        try:
            from utils.checklist_printer import ChecklistPrinter

            printer = ChecklistPrinter()
            if printer.print_dialog(self.root_items, self):
                self.status_bar.showMessage("체크리스트를 프린트했습니다.")
            else:
                self.status_bar.showMessage("프린트가 취소되었습니다.")

        except ImportError as e:
            QMessageBox.critical(self, '모듈 오류', f'프린트 모듈을 불러올 수 없습니다:\n{str(e)}')
        except Exception as e:
            QMessageBox.critical(self, '오류', f'프린트 중 오류가 발생했습니다:\n{str(e)}')

    def export_checklist_to_pdf(self):
        """체크리스트를 PDF로 저장"""
        if not self.root_items:
            QMessageBox.information(self, 'PDF 저장 불가', '저장할 체크리스트가 없습니다.')
            return

        filename, _ = QFileDialog.getSaveFileName(
            self,
            'PDF로 저장',
            'checklist.pdf',
            'PDF Files (*.pdf);;All Files (*)'
        )

        if filename:
            try:
                from utils.checklist_printer import ChecklistPrinter

                printer = ChecklistPrinter()
                printer.export_to_pdf(self.root_items, filename)
                self.status_bar.showMessage(f"PDF로 저장 완료: {filename}")
                QMessageBox.information(self, 'PDF 저장 완료', f'체크리스트가 PDF로 저장되었습니다:\n{filename}')

            except ImportError as e:
                QMessageBox.critical(self, '모듈 오류', f'프린트 모듈을 불러올 수 없습니다:\n{str(e)}')
            except Exception as e:
                QMessageBox.critical(self, '오류', f'PDF 저장 중 오류가 발생했습니다:\n{str(e)}')

    def closeEvent(self, event):
        if self.is_modified:
            reply = QMessageBox.question(self, '종료',
                                       '변경사항이 있습니다. 저장하지 않고 종료하시겠습니까?',
                                       QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            if reply != QMessageBox.Yes:
                event.ignore()
                return

        # 프로그램 종료 시 모든 파일 잠금 해제
        self.unlock_all_files()
        event.accept()

    def save_current_as_training_data(self):
        """현재 작업을 학습 데이터로 저장"""
        if not self.training_collector:
            QMessageBox.warning(self, '기능 없음', '학습 데이터 수집 기능을 사용할 수 없습니다.')
            return

        if not self.selected_widget:
            QMessageBox.information(self, '선택 필요', '학습 데이터로 저장할 토글을 선택해주세요.')
            return

        try:
            # 선택된 토글을 학습 데이터로 저장
            toggle_item = self.selected_widget.item
            # ToggleItem의 source_file 속성을 사용 (PDF 경로)
            original_pdf_path = getattr(toggle_item, 'source_file', None)
            self.training_collector.save_toggle_as_training_data(toggle_item, original_pdf_path)

            QMessageBox.information(
                self,
                '저장 완료',
                f'학습 데이터가 저장되었습니다!\n\n'
                f'파일: {self.training_collector.data_file}'
            )
            self.status_bar.showMessage('학습 데이터 저장 완료')

        except Exception as e:
            QMessageBox.critical(self, '저장 오류', f'학습 데이터 저장 중 오류가 발생했습니다:\n{str(e)}')

    def show_training_statistics(self):
        """학습 데이터 통계 표시"""
        if not self.training_collector:
            return

        stats = self.training_collector.get_statistics()

        stats_text = f"📊 학습 데이터 통계\n\n"
        stats_text += f"총 수정 횟수: {stats['total_corrections']}개\n\n"

        if stats['files']:
            stats_text += "파일별 통계:\n"
            for file_info in stats['files']:
                stats_text += f"  • {file_info['name']}: {file_info['count']}개\n"
        else:
            stats_text += "아직 저장된 학습 데이터가 없습니다.\n"

        if stats['latest']:
            stats_text += f"\n최근 저장: {stats['latest']['timestamp']}"

        QMessageBox.information(self, '학습 데이터 통계', stats_text)

    def export_openai_format(self):
        """OpenAI 파인튜닝 형식으로 내보내기"""
        if not self.training_collector:
            return

        try:
            output_file = self.training_collector.export_for_openai_finetuning()
            if output_file:
                QMessageBox.information(
                    self,
                    '내보내기 완료',
                    f'OpenAI 파인튜닝 파일이 생성되었습니다!\n\n{output_file}'
                )
            else:
                QMessageBox.warning(self, '내보내기 실패', '저장된 학습 데이터가 없습니다.')
        except Exception as e:
            QMessageBox.critical(self, '내보내기 오류', f'파일 생성 중 오류가 발생했습니다:\n{str(e)}')

    def export_local_format(self):
        """로컬 파인튜닝 형식으로 내보내기"""
        if not self.training_collector:
            return

        try:
            output_file = self.training_collector.export_for_local_finetuning()
            if output_file:
                QMessageBox.information(
                    self,
                    '내보내기 완료',
                    f'로컬 파인튜닝 파일이 생성되었습니다!\n\n{output_file}'
                )
            else:
                QMessageBox.warning(self, '내보내기 실패', '저장된 학습 데이터가 없습니다.')
        except Exception as e:
            QMessageBox.critical(self, '내보내기 오류', f'파일 생성 중 오류가 발생했습니다:\n{str(e)}')


def main():
    app = QApplication(sys.argv)
    
    app.setStyle('Fusion')
    
    window = RiskManagementApp()
    window.show()
    
    sys.exit(app.exec_())


if __name__ == '__main__':
    main()